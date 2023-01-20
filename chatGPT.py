# Gerado pelo chatGPT: 'Este código importa as bibliotecas csv e openpyxl, cria uma nova planilha de trabalho (Workbook) e define a planilha ativa como "at02_cache". 
# Ele então abre um arquivo CSV chamado "AT-02 Quantidade de Pacientes e Procedimentos por Estabelecimento por mês.csv" e adiciona cada linha 
# a "at02_cache". Ele remove as primeiras três linhas de "at02_cache" e modifica uma célula específica se o valor da célula contiver 11 caracteres. 
# Ele carrega uma planilha de trabalho chamada "DATA.xlsx" e define as variáveis "unidades", "procs" e "non_procs" como suas respectivas 
# planilhas. Ele cria um dicionário vazio chamado "v_unid_nao_pertence" e outro chamado "v_unidades", e então preenche "v_unidades" com dados 
# das células da planilha "unidades". Ele define uma função chamada "extract_nonproc" que extrai dados de uma planilha específica e armazena-os 
# em um dicionário. Ele cria uma nova planilha chamada "AT-02 Quantidade de Pacientes e" e, em seguida, itera sobre cada linha de "at02_cache" e 
# verifica se determinadas condições são verdadeiras antes de adicionar a linha à nova planilha.'

#Importa as bibliotecas
import csv #biblioteca do sistema
import openpyxl #biblioteca instalada pelo PIP
from openpyxl import load_workbook

wb = openpyxl.Workbook() #cria um workbook, uma planilha com uma aba
at02_cache = wb.active  #cria variável da aba recém criada

#Abre o arquivo CSV e converte para o formato excel na aba recém criada, renomeie aqui no formato 'arquivo.csv'
with open('AT-02 Quantidade de Pacientes e Procedimentos por Estabelecimento por mês.csv', encoding="utf8") as f: 
    reader = csv.reader(f, delimiter = ',')
    for row in reader:
        at02_cache.append(row)

at02_cache.delete_rows(1,3) #deleta o cabeçalho do arquivo CSV, da 1ª até a 3ª linha

#Essa função trata a coluna N (procedimentos), cortando os dois últimos digítos com [:-2]
for row in at02_cache.iter_rows(min_row=2, max_row=at02_cache.max_row):
    if len(row[13].value) == 11:
        at02_cache.cell(row=row[13].row, column=14).value = row[13].value[:-2]

#Carregando o workbook DATA.xlsx e carregando as planilhas
data = load_workbook("DATA.xlsx")
unidades = data["UNIDADES"]
procs = data["PROCS"]
non_procs = data["NON_PROCS"]

#Criação de dicionários
v_unid_nao_pertence = {} # armazenamento das unidades que não pertencem ao contrato
v_unidades = {} #unidades do contrato como constam em ambas as planilhas, CMES e categoria (AMA, UBS, CAPS, etc)

#Cria-se uma tupla de uma linha selecionada (N, i) para servir de chave de um dicionário
chaves_unidades = [unidades.cell(1, i).value for i in range(1, unidades.max_column + 1)]

#Itera a planilha 'unidades', preenchendo os dicionários 'v_unidades' e 'v_unid_nao_pertence' conforme critério
for row in unidades.iter_rows(min_row=2, values_only=True):
    # caso a coluna D 'CATEGORIA' não tenha o valor "NAO PERTENCE", ele popula o dicionário v_unidades
    if row[3] != "NAO_PERTENCE":
        for i in range(len(row)):
            if chaves_unidades[i] not in v_unidades: # verifica se a chave correspondente ao índice atual do loop "for" já existe no dicionário "v_unidades"
                v_unidades[chaves_unidades[i]] = [] #popula as chaves com valores originais da variavel 'chaves_unidades' 
            v_unidades[chaves_unidades[i]].append(row[i])
    # senão será populado o dicionário das unidades de fora do contrato para posterior filtragem
    else:
        v_unid_nao_pertence[row[1]] = row[3]

#Essa função processa a planilha 'NON_PROCS' em dicionários com chaves aninhadas que precisam de função recursiva para chegar a listas de valores
def extract_nonproc(planilha):
    headers = [cell.value for cell in non_procs[1]] #cria chaves a partir do cabeçalho linha 1 da planilha
    data = {}
    current_key = None 
    for row in range(2, non_procs.max_row + 1): #itera por todas as linhas, menos o cabeçalho
        key = non_procs.cell(row, 2).value #define a chave da linha como o valor da coluna B 'PRODUÇÃO' exportada do webssas
        if key != current_key: #compara a chave atual com a chave da linha anterior
            current_key = key #caso a chave tenha mudado, estabele a nova chave para futura comparação
            data[key] = {} #cria um novo dicionário para a nova chave
            for col in range(4, 10): #itera as colunas D:I e popula o dicionário com chaves 'headers'
                data[key][headers[col-1]] = [] #'col - 1' porque o python trabalha com índice[0] enquanto o excel tem colunas com índice a partir de [1]
        for col in range(4, 10): #popula as chaves do dicionário atual com valores de cada linha
            data[key][headers[col-1]].append(non_procs.cell(row, col).value)
    return data

v_non_proc = extract_nonproc(non_procs)

#função recursiva que busca chaves dentro de chaves, até conseguir acessar a lista de valores
def iterar_v_non_proc(dicionario):
    for k, valor in dicionario.items(): #enumera as chaves
        if isinstance(valor, dict): #Se a chave conter outras chaves e não uma lista de valores a função é chamada recursivamente passando a chave aninhada como parâmetro
            result = iterar_v_non_proc(valor) #a função é armazenada na variável 'result' e retornada para que o valor não se perca
            if result: # Se o resultado é verdadeiro, interrompe a função e retorna o resultado
                return result
        elif k == 'Nome_CBO1': #uma vez encontrada as chaves com acesso direto aos valores, ele procura a chave de interesse 'Nome_CBO1'
            for i, nome_cbo in enumerate(valor): #itera sobre a tupla de valores do CBO de não procedimentos
                if nome_cbo == nome_cbo_at02: #encontra o valor do CBO da linha AT-02 processada dentro da lista de não procedimentos
                    if dicionario['Código_Procedimento'][i] == cod_proc_at02: #dentre todos os procedimentos para o CBO encontrado, 
                    #ele compara os não-procedimentos com o procedimento atual até encontra-lo, iterando pelo índice [i] do loop FOR
                        if dicionario['CATEGORIA'][i] == None: 
                            return False #não havendo categoria no não-procedimento, ele será descartado
                        elif dicionario['CATEGORIA'][i] == categoria_estab_at02:
                            return False #compara a categoria da unidade em at-02 com a do não-procedimento do CBO e Procedimento encontrados e então descarta
            return True #não havendo não-procedimento encontrado, mantém o valor na planilha

# Deletar linhas de uma planilha com muitas linhas é um processo lento, pois o python atualiza toda a planilha a cada linha excluída
# Um método melhor é criar uma nova planilha e copiar apenas as linhas necessárias, e depois excluir a planilha original
at02 = wb.create_sheet("AT-02 Quantidade de Pacientes e")

for row in at02_cache.iter_rows(min_row=1, max_row=at02_cache.max_row, values_only=True):
    append_at02 = True #A principio, todas as linhas são copiadas
    nome_estabelecimento_at02 = row[5] #essa lista de variáveis são os valores que serão usados para comparação e validação
    categoria_estab_at02 = None
    nome_cbo_at02 = row[11]
    cod_proc_at02 = row[13]
    for i, valores in enumerate(v_unidades['H1___Nome_Estabelecimento2']): #compara o estabelecimento em AT-02 e encontra na lista 'unidades' do data
        if valores == nome_estabelecimento_at02:
            categoria_estab_at02 = v_unidades['CATEGORIA'][i] #quando o valor é encontrado, copia 'CATEGORIA' da planilha 'unidades'
            break
    append_at02 = iterar_v_non_proc(v_non_proc) #chama a função que itera o dicionário de não procedimentos
    if nome_estabelecimento_at02 in v_unid_nao_pertence: #checa se a unidade está dentro do contrato
        append_at02 = False
    # if nome_cbo_at02 != 'Medico Pediatra': #testar CBO, manter a estrutura comentada
    #     append_at02 = False
    if append_at02 == True: 
        at02.append(row)

wb.remove(at02_cache) #exclui a aba de planilha antes do filtro
wb.save('AT-02 Quantidade de Pacientes e Procedimentos por Estabelecimento por mês.xlsx')