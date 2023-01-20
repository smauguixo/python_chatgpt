# Esse script transforma o bando de dados AT-02 em formato CSV numa planilha de excel, filtrando as informações que não serão usadas

#importa as bibliotecas
import csv #biblioteca do sistema
import openpyxl #biblioteca instalada pelo PIP

contRow = 1 #contador de linhas

wb = openpyxl.Workbook() #cria uma planilha com uma aba
ws = wb.active  #cria uma variável para a aba ativa da planilha recém criada
ws.title = "AT-02" #nomeia a aba

#abre o arquivo CSV e converte em colunas, renomeie aqui no formato 'arquivo.csv' 
with open('AT-02 Quantidade de Pacientes e Procedimentos por Estabelecimento por mês.csv', encoding="utf8") as f: 
    reader = csv.reader(f, delimiter = ',')
    for row in reader:
        ws.append(row)

ws.delete_rows(1,3) #deleta o cabeçalho do arquivo CSV, o parâmetro atual é da 1ª até a 3ª linha >> (1,3)

#essa rotina trata a coluna N, cortando o último dígito do código do procedimento. Caso a célula tenha 11 dígitos, ele corta os dois últimos.
while True:
    if ws['N' + str(contRow)].value is not None :
        if len(ws['N' + str(contRow)].value) == 11 :
            ws['N' + str(contRow)].value = ws['N' + str(contRow)].value[:-2]
    else:
        break
    contRow+=1   

contRow = 1 #limpa o contador

#cria uma nova aba e filtra as unidades que não fazem parte do contrato no momento
ws1 = wb.create_sheet('ws1') 
for row in ws.values:
    if ws['F' + str(contRow)].value != ("unidade blablabla 1") : #copia e cola o valor da célula do excel 
        if ws['F' + str(contRow)].value != ("unidade blablabla 2") :
            if ws['F' + str(contRow)].value != ("unidade blablabla 3") :
                if ws['F' + str(contRow)].value != ("unidade blablabla 4") :
                    if ws['F' + str(contRow)].value != ("unidade blablabla 5") :
                        if ws['F' + str(contRow)].value != ("unidade blablabla 6") :
                            if ws['F' + str(contRow)].value != ("unidade blablabla 7") :
                                if ws['F' + str(contRow)].value != ("unidade blablabla 8") :
                                    ws1.append(row)
    contRow+=1

wb.remove(wb["AT-02"]) #exclui a aba de planilha antes do filtro
contRow = 1 #limpa o contador

#cria uma nova aba e filtra as unidades que não fazem parte do contrato no momento
ws2 = wb.create_sheet('AT-02 Quantidade de Pacientes e') 
for row in ws1.values:
    if ws1['L' + str(contRow)].value != ("Medico Pediatra"):
        ws2.append(row)
    if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("030106003"): #copia e cola o valor da célula do excel 
        if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("030106004"):
            if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("030106005"):
                if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("010104002"):
                    if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("030101026"):
                        if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("030101027"):
                            if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("030101028"):
                                if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("040101006"):
                                    if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("021102003"):
                                        if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("010104007"):
                                            if ws1['L' + str(contRow)].value == ("Medico Pediatra") and ws1['N' + str(contRow)].value != ("010104008"):
                                                ws2.append(row)
                                            
    contRow+=1
   

wb.remove(wb["ws1"]) #exclui a aba de planilha antes do filtro
wb.save('AT-02 Quantidade de Pacientes e Procedimentos por Estabelecimento por mês.xlsx')
